import openpyxl

def getRowCount(PathExcelFile, SheetName):
    WorkBook = openpyxl.load_workbook(PathExcelFile)
    WorkSheet = WorkBook.get_sheet_by_name(SheetName)
    return(WorkSheet.max_row)

def getColumnCount(PathExcelFile, SheetName):
    WorkBook = openpyxl.load_workbook(PathExcelFile)
    WorkSheet = WorkBook.get_sheet_by_name(SheetName)
    return(WorkSheet.max_column)

def readData(PathExcelFile, SheetName, RowNum, ColumnNum):
    WorkBook = openpyxl.load_workbook(PathExcelFile)
    WorkSheet = WorkBook.get_sheet_by_name(SheetName)
    return WorkSheet.cell(row=RowNum, column=ColumnNum).value

def writeData(PathExcelFile, SheetName, RowNum, ColumnNum, data):
    WorkBook = openpyxl.load_workbook(PathExcelFile)
    WorkSheet = WorkBook.get_sheet_by_name(SheetName)
    WorkSheet.cell(row=RowNum, column=ColumnNum).value = data
    WorkBook.save(PathExcelFile)